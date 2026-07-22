import calendar
import os
import subprocess
import sys
import traceback
import pandas as pd
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# -- Paleta de colores (identidad morada, a juego con el icono y el Excel) --
BG_APP        = "#F4F2F8"   # fondo general, lavanda muy claro
BG_CARD       = "#FFFFFF"   # fondo de tarjetas
BG_HEADER     = "#3B2B4F"   # barra de título, morado oscuro
FG_HEADER     = "#FFFFFF"
FG_HEADER_SUB = "#C8B9DC"   # subtítulo de la cabecera
FG_TITLE      = "#241A31"
FG_LABEL      = "#3A3A3C"
FG_MUTED      = "#8E8496"
COLOR_ACCENT  = "#6C4C87"
COLOR_HOVER   = "#5A3E70"
COLOR_PRESS   = "#4E3366"
COLOR_DISABLED = "#B7A8C6"
COLOR_BORDER  = "#DCD5E6"
COLOR_CHIP_BG = "#E9E2F1"   # fondo del chip de preview

# -- Excel --
COLOR_HEADER_XL = "6C4C87"
COLOR_ROW_ODD   = "C2B4CA"
COLOR_ROW_EVN   = "DCD3E1"

# ---------------------------------------------------------------------- #
#  CONFIGURACIÓN DEL EQUIPO Y TAREAS                                      #
#  (único sitio a editar si cambia el personal o su disponibilidad)       #
# ---------------------------------------------------------------------- #

# Solo se limpia de lunes a viernes (no hay limpieza en sábado, domingo
# ni en los turnos de noche de los AP).
DIAS_SEMANA = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]

# Zonas de Juan — las limpian los AP, cada uno solo en los días que trabaja
# (turnos de día de lunes a viernes; los turnos de noche no limpian).
ESTANCIAS_JUAN = ["Hab. Juan", "Hab. AP", "Baño Juan"]
STAFF_JUAN = ["Dora", "Eva", "Lina", "Miguel", "Sandra", "Valentina"]
DIAS_DISPONIBLES = {
    "Dora":      ["Lunes"],
    "Eva":       ["Miércoles", "Jueves", "Viernes"],
    "Lina":      ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"],
    "Miguel":    ["Lunes", "Martes"],
    "Sandra":    ["Martes"],
    "Valentina": ["Martes", "Miércoles", "Jueves", "Viernes"],
}
# Tareas comunes dentro de Zonas de Juan, rotando en ciclos de 3 semanas.
TAREAS_COMUNES_JUAN = {
    0: [("Miércoles", "Entrada")],
    1: [("Viernes",   "Cocina")],
    2: [("Lunes",     "Salón"), ("Martes", "Cocina")],
}

# Zonas Comunes
STAFF_COMUNES = ["Lina", "Angie", "Juan"]
# Cada bloque es una lista de (día, tarea): la persona que cubre el bloque
# hace todas esas tareas, en esos días fijos (solo de lunes a viernes).
BLOQUES_COMUNES = [
    [("Lunes",     "Salón"),   ("Martes",  "Cocina")],
    [("Miércoles", "Entrada"), ("Jueves",  "Salón")],
    [("Viernes",   "Cocina")],
]


def _card(parent, title=""):
    """Frame tipo tarjeta con borde sutil y esquinas simuladas."""
    outer = tk.Frame(parent, bg=COLOR_BORDER, padx=1, pady=1)
    inner = tk.Frame(outer, bg=BG_CARD, padx=18, pady=14)
    inner.pack(fill="both", expand=True)
    if title:
        tk.Label(
            inner, text=title,
            font=("Helvetica", 10, "bold"),
            bg=BG_CARD, fg=FG_MUTED
        ).pack(anchor="w", pady=(0, 8))
    return outer, inner


class RoundButton(tk.Canvas):
    """Botón con esquinas redondeadas, hover y estado deshabilitado."""

    def __init__(self, master, text, command, height=48, radius=14,
                 font=("Helvetica", 13, "bold")):
        super().__init__(master, height=height, bg=master["bg"],
                         highlightthickness=0, cursor="hand2")
        self._command = command
        self._text = text
        self._font = font
        self._radius = radius
        self._state = "normal"
        self._fill = COLOR_ACCENT

        self.bind("<Configure>", lambda _: self._draw())
        self.bind("<Enter>", lambda _: self._set_fill(COLOR_HOVER))
        self.bind("<Leave>", lambda _: self._set_fill(COLOR_ACCENT))
        self.bind("<Button-1>", lambda _: self._set_fill(COLOR_PRESS))
        self.bind("<ButtonRelease-1>", self._on_release)

    def _set_fill(self, color):
        if self._state == "normal":
            self._fill = color
            self._draw()

    def _on_release(self, event):
        if (self._state == "normal"
                and 0 <= event.x <= self.winfo_width()
                and 0 <= event.y <= self.winfo_height()):
            self._set_fill(COLOR_HOVER)
            self._command()

    def _draw(self):
        self.delete("all")
        w, h = self.winfo_width(), self.winfo_height()
        if w <= 1 or h <= 1:
            return
        r = self._radius
        fill = COLOR_DISABLED if self._state == "disabled" else self._fill
        pts = [
            r, 0, w - r, 0, w, 0, w, r, w, h - r, w, h,
            w - r, h, r, h, 0, h, 0, h - r, 0, r, 0, 0,
        ]
        self.create_polygon(pts, smooth=True, fill=fill)
        self.create_text(w // 2, h // 2, text=self._text,
                         fill="#FFFFFF", font=self._font)

    def config(self, cnf=None, **kw):
        state = kw.pop("state", None)
        text = kw.pop("text", None)
        if state is not None:
            self._state = "disabled" if str(state) == "disabled" else "normal"
            self._fill = COLOR_ACCENT
            super().config(cursor="arrow" if self._state == "disabled" else "hand2")
        if text is not None:
            self._text = text
        if state is not None or text is not None:
            self._draw()
        if cnf or kw:
            return super().config(cnf, **kw) if cnf else super().config(**kw)

    configure = config


class AppTurnosNativa:
    def __init__(self, root):
        self.root = root
        self.root.title("Gestor de Turnos")
        self.root.resizable(False, False)
        self.root.configure(bg=BG_APP)

        self._build_header()
        self._build_body()
        self._centrar_ventana(560, 330)

    # ------------------------------------------------------------------ #
    #  UI BUILD                                                            #
    # ------------------------------------------------------------------ #

    def _build_header(self):
        hdr = tk.Frame(self.root, bg=BG_HEADER, height=72)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(
            hdr, text="✦  Gestor de Turnos",
            font=("Helvetica", 16, "bold"),
            bg=BG_HEADER, fg=FG_HEADER
        ).place(relx=0.5, rely=0.36, anchor="center")
        tk.Label(
            hdr, text="Planificador de limpieza semanal",
            font=("Helvetica", 10),
            bg=BG_HEADER, fg=FG_HEADER_SUB
        ).place(relx=0.5, rely=0.70, anchor="center")

    def _build_body(self):
        body = tk.Frame(self.root, bg=BG_APP)
        body.pack(fill="both", expand=True, padx=24, pady=20)

        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "Date.TCombobox",
            fieldbackground=BG_CARD,
            background=BG_CARD,
            foreground=FG_TITLE,
            arrowcolor=COLOR_ACCENT,
            bordercolor=COLOR_BORDER,
            lightcolor=BG_CARD,
            darkcolor=BG_CARD,
            selectbackground=BG_CARD,
            selectforeground=FG_TITLE,
            relief="flat",
            padding=3,
        )
        style.map("Date.TCombobox",
                  fieldbackground=[("readonly", BG_CARD)],
                  foreground=[("readonly", FG_TITLE)])
        self.root.option_add("*TCombobox*Listbox.selectBackground", COLOR_ACCENT)
        self.root.option_add("*TCombobox*Listbox.selectForeground", "#FFFFFF")

        # Defaults dinámicos
        today = datetime.today()
        m_end = today.month + 2
        y_end = today.year
        if m_end > 12:
            m_end -= 12
            y_end += 1
        last_day_end = calendar.monthrange(y_end, m_end)[1]

        # --- Tarjetas de fecha, lado a lado ---
        cards_row = tk.Frame(body, bg=BG_APP)
        cards_row.pack(fill="x")
        cards_row.columnconfigure(0, weight=1, uniform="fechas")
        cards_row.columnconfigure(2, weight=1, uniform="fechas")

        card_outer, card_ini = _card(cards_row, "DESDE")
        card_outer.grid(row=0, column=0, sticky="nsew")

        tk.Label(
            cards_row, text="→", font=("Helvetica", 15),
            bg=BG_APP, fg=FG_MUTED
        ).grid(row=0, column=1, padx=8)

        card_outer2, card_fin = _card(cards_row, "HASTA")
        card_outer2.grid(row=0, column=2, sticky="nsew")

        row_ini = tk.Frame(card_ini, bg=BG_CARD)
        row_ini.pack(anchor="w")
        self.d_ini, self.m_ini, self.a_ini = self._date_row(
            row_ini,
            str(today.day).zfill(2),
            str(today.month).zfill(2),
            str(today.year),
        )

        row_fin = tk.Frame(card_fin, bg=BG_CARD)
        row_fin.pack(anchor="w")
        self.d_fin, self.m_fin, self.a_fin = self._date_row(
            row_fin,
            str(last_day_end).zfill(2),
            str(m_end).zfill(2),
            str(y_end),
        )

        # --- Chip de preview de semanas ---
        self._preview_var = tk.StringVar(value="")
        tk.Label(
            body,
            textvariable=self._preview_var,
            font=("Helvetica", 11, "bold"),
            bg=COLOR_CHIP_BG, fg=COLOR_ACCENT,
            padx=16, pady=6,
        ).pack(pady=(18, 18))

        # --- Botón principal ---
        self.btn = RoundButton(
            body,
            text="Generar y Guardar Excel",
            command=self.procesar,
        )
        self.btn.pack(fill="x")

        # Bind: ajustar días válidos del mes y refrescar preview
        for widget in (self.d_ini, self.m_ini, self.a_ini,
                        self.d_fin, self.m_fin, self.a_fin):
            widget.bind("<<ComboboxSelected>>", lambda _: self._on_fecha_cambiada())

        self._on_fecha_cambiada()

    def _date_row(self, parent, day, month, year):
        """Fila compacta DD / MM / AAAA con estilo coherente."""
        font_combo = ("Helvetica", 13)
        combo_day   = ttk.Combobox(parent, values=[str(i).zfill(2) for i in range(1, 32)],
                                    width=3, state="readonly", style="Date.TCombobox",
                                    font=font_combo, justify="center")
        combo_month = ttk.Combobox(parent, values=[str(i).zfill(2) for i in range(1, 13)],
                                    width=3, state="readonly", style="Date.TCombobox",
                                    font=font_combo, justify="center")
        combo_year  = ttk.Combobox(parent, values=[str(i) for i in range(2024, 2036)],
                                    width=5, state="readonly", style="Date.TCombobox",
                                    font=font_combo, justify="center")

        combo_day.set(day)
        combo_month.set(month)
        combo_year.set(year)

        sep_kw = dict(bg=BG_CARD, fg=FG_MUTED, font=("Helvetica", 13))

        combo_day.pack(side="left", padx=(0, 2))
        tk.Label(parent, text="/", **sep_kw).pack(side="left")
        combo_month.pack(side="left", padx=(2, 2))
        tk.Label(parent, text="/", **sep_kw).pack(side="left")
        combo_year.pack(side="left", padx=(2, 0))

        return combo_day, combo_month, combo_year

    # ------------------------------------------------------------------ #
    #  HELPERS                                                             #
    # ------------------------------------------------------------------ #

    def _centrar_ventana(self, ancho, alto):
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth()  - ancho) // 2
        y = (self.root.winfo_screenheight() - alto)  // 2
        self.root.geometry(f"{ancho}x{alto}+{x}+{y}")

    def _on_fecha_cambiada(self):
        self._ajustar_dias_del_mes(self.d_ini, self.m_ini, self.a_ini)
        self._ajustar_dias_del_mes(self.d_fin, self.m_fin, self.a_fin)
        self._actualizar_preview()

    def _ajustar_dias_del_mes(self, combo_d, combo_m, combo_a):
        """Limita el combo de día a los días reales del mes/año elegidos."""
        try:
            ultimo = calendar.monthrange(int(combo_a.get()), int(combo_m.get()))[1]
        except ValueError:
            return
        combo_d["values"] = [str(i).zfill(2) for i in range(1, ultimo + 1)]
        if int(combo_d.get()) > ultimo:
            combo_d.set(str(ultimo).zfill(2))

    def _actualizar_preview(self):
        inicio = self.obtener_fecha(self.d_ini.get(), self.m_ini.get(), self.a_ini.get())
        fin    = self.obtener_fecha(self.d_fin.get(), self.m_fin.get(), self.a_fin.get())
        if inicio and fin and fin > inicio:
            n = len(self._calcular_semanas(inicio, fin))
            self._preview_var.set(f"{n} semana{'s' if n != 1 else ''} a generar")
        else:
            self._preview_var.set("Comprueba las fechas")

    def obtener_fecha(self, d, m, a):
        try:
            return datetime(int(a), int(m), int(d))
        except ValueError:
            return None

    def _calcular_semanas(self, inicio, fin):
        """Semanas (lunes-domingo) que cubren el rango [inicio, fin]."""
        curr = inicio - timedelta(days=inicio.weekday())
        semanas = []
        while curr <= fin:
            semanas.append((curr, curr + timedelta(days=6)))
            curr += timedelta(days=7)
        return semanas

    # ------------------------------------------------------------------ #
    #  LÓGICA DE TURNOS                                                    #
    # ------------------------------------------------------------------ #

    def generar_juan(self, semanas):
        estancias = ESTANCIAS_JUAN
        staff = STAFF_JUAN
        dias_disponibles = DIAS_DISPONIBLES

        rows = []
        ocupados_por_semana = []
        idx = 0
        idx_comunes = 0
        # Puntero de rotación por persona para no repetir siempre su primer día disponible.
        rotacion_dia = {persona: 0 for persona in staff}

        for i, (f_ini, f_fin) in enumerate(semanas):
            row = {"Semana": f"{f_ini.strftime('%d/%m')} -\n{f_fin.strftime('%d/%m')}"}
            row.update({d: "-" for d in DIAS_SEMANA})

            # Días ya usados esta semana por cada persona, compartido entre
            # el reparto de estancias y el de tareas comunes para no
            # apilar dos tareas de la misma persona en un mismo día.
            personas_por_dia = {d: set() for d in DIAS_SEMANA}

            def _asignar(persona, tarea, dia):
                etiqueta = f"{tarea} ({persona})"
                if row[dia] == "-":
                    row[dia] = etiqueta
                else:
                    row[dia] += f"\n| {etiqueta}"
                personas_por_dia[dia].add(persona)

            n_staff = len(staff)
            p1 = staff[idx % n_staff]
            p2 = staff[(idx + 1) % n_staff]
            p3 = staff[(idx + 2) % n_staff]
            idx += 3

            asignaciones = [(p1, estancias[0]), (p2, estancias[1]), (p3, estancias[2])]

            for persona, tarea in asignaciones:
                dias_p = dias_disponibles[persona]
                n = len(dias_p)
                inicio_rot = rotacion_dia[persona]

                dia_elegido = None
                for offset in range(n):
                    candidato_dia = dias_p[(inicio_rot + offset) % n]
                    if not personas_por_dia[candidato_dia]:
                        dia_elegido = candidato_dia
                        break
                if dia_elegido is None:
                    dia_elegido = dias_p[inicio_rot % n]

                rotacion_dia[persona] = (dias_p.index(dia_elegido) + 1) % n
                _asignar(persona, tarea, dia_elegido)

            juan_comunes_tareas = TAREAS_COMUNES_JUAN[i % len(TAREAS_COMUNES_JUAN)]

            for dia, tarea in juan_comunes_tareas:
                candidato_libre = candidato_ocupado = None
                offset_libre = offset_ocupado = None

                for offset in range(len(staff)):
                    candidato = staff[(idx_comunes + offset) % len(staff)]
                    if dia not in dias_disponibles[candidato]:
                        continue
                    if candidato not in personas_por_dia[dia]:
                        candidato_libre, offset_libre = candidato, offset
                        break
                    elif candidato_ocupado is None:
                        candidato_ocupado, offset_ocupado = candidato, offset

                if candidato_libre is not None:
                    candidato, offset = candidato_libre, offset_libre
                elif candidato_ocupado is not None:
                    candidato, offset = candidato_ocupado, offset_ocupado
                else:
                    raise ValueError(
                        f"Nadie del equipo está disponible el {dia} para la tarea "
                        f"'{tarea}'. Revisa 'dias_disponibles'."
                    )

                _asignar(candidato, tarea, dia)
                idx_comunes = (idx_comunes + offset + 1) % len(staff)

            ocupados_por_semana.append(personas_por_dia)
            rows.append(row)

        return pd.DataFrame(rows), ocupados_por_semana

    def generar_comunes(self, semanas, ocupados_juan=None):
        staff = STAFF_COMUNES
        bloques = BLOQUES_COMUNES
        ocupados_juan = ocupados_juan or [{}] * len(semanas)

        def _dias(bloque):
            return [dia for dia, _ in bloque]

        def _ocupado(persona, ocupado_semana, bloque):
            return any(persona in ocupado_semana.get(dia, ()) for dia in _dias(bloque))

        rows = []
        n_staff = len(staff)
        for i, (f_ini, f_fin) in enumerate(semanas):
            personas = [staff[(i + k) % n_staff] for k in range(len(bloques))]
            ocupado_semana = ocupados_juan[i] if i < len(ocupados_juan) else {}

            # Si a alguien ya le toca una tarea de "Zonas de Juan" un día de
            # su bloque (p.ej. Lina, que está en ambos equipos), se
            # intercambia su bloque con el de otra persona libre esos días.
            for b_idx, bloque in enumerate(bloques):
                if not _ocupado(personas[b_idx], ocupado_semana, bloque):
                    continue
                for j in range(len(personas)):
                    if j == b_idx:
                        continue
                    if (not _ocupado(personas[j], ocupado_semana, bloque)
                            and not _ocupado(personas[b_idx], ocupado_semana, bloques[j])):
                        personas[b_idx], personas[j] = personas[j], personas[b_idx]
                        break

            row = {"Semana": f"{f_ini.strftime('%d/%m')} -\n{f_fin.strftime('%d/%m')}"}
            row.update({d: "-" for d in DIAS_SEMANA})
            for b_idx, bloque in enumerate(bloques):
                persona = personas[b_idx]
                for dia, tarea in bloque:
                    row[dia] = f"{tarea} ({persona})"
            rows.append(row)

        return pd.DataFrame(rows)

    def aplicar_estilos_excel(self, ruta, df_comunes, df_juan):
        with pd.ExcelWriter(ruta, engine='openpyxl') as writer:
            df_comunes.to_excel(writer, sheet_name='Zonas Comunes', index=False)
            df_juan.to_excel(writer, sheet_name='Zonas de Juan', index=False)

            workbook = writer.book

            fill_header = PatternFill(start_color=COLOR_HEADER_XL, end_color=COLOR_HEADER_XL, fill_type="solid")
            fill_odd    = PatternFill(start_color=COLOR_ROW_ODD,   end_color=COLOR_ROW_ODD,   fill_type="solid")
            fill_even   = PatternFill(start_color=COLOR_ROW_EVN,   end_color=COLOR_ROW_EVN,   fill_type="solid")

            font_header = Font(color="FFFFFF", bold=True,  size=12)
            font_normal = Font(color="000000", bold=False, size=12)

            align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

            white_side   = Side(border_style="thin", color="FFFFFF")
            border_white = Border(left=white_side, right=white_side, top=white_side, bottom=white_side)

            for sheet_name in ['Zonas Comunes', 'Zonas de Juan']:
                ws = workbook[sheet_name]
                ws.freeze_panes = "B2"

                for col_idx in range(1, ws.max_column + 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 20

                for row_idx, row in enumerate(ws.iter_rows()):
                    # Alto según la celda con más líneas apiladas de la fila.
                    max_lineas = max(
                        str(cell.value).count("\n") + 1 if cell.value else 1
                        for cell in row
                    )
                    ws.row_dimensions[row[0].row].height = max(45, 22 * max_lineas + 8)

                    for cell in row:
                        cell.alignment = align_center
                        cell.border = border_white

                        if row_idx == 0 or cell.column == 1:
                            cell.fill = fill_header
                            cell.font = font_header
                        else:
                            cell.font = font_normal
                            cell.fill = fill_odd if row_idx % 2 != 0 else fill_even

    def procesar(self):
        d_i, m_i, a_i = self.d_ini.get(), self.m_ini.get(), self.a_ini.get()
        d_f, m_f, a_f = self.d_fin.get(), self.m_fin.get(), self.a_fin.get()

        inicio = self.obtener_fecha(d_i, m_i, a_i)
        fin    = self.obtener_fecha(d_f, m_f, a_f)

        if not inicio or not fin:
            messagebox.showerror("Error", "La fecha seleccionada no existe.")
            return
        if fin <= inicio:
            messagebox.showerror("Error", "La fecha de fin debe ser posterior a la de inicio.")
            return

        semanas = self._calcular_semanas(inicio, fin)

        nombre_archivo_sugerido = (
            f"Turnos limpieza {inicio.strftime('%d-%m-%Y')} a {fin.strftime('%d-%m-%Y')}.xlsx"
        )

        ruta = filedialog.asksaveasfilename(
            initialfile=nombre_archivo_sugerido,
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )

        if not ruta:
            return

        self.btn.config(state="disabled", text="Generando…")
        self.root.config(cursor="watch")
        self.root.update()

        try:
            df_j, ocupados_juan = self.generar_juan(semanas)
            df_c = self.generar_comunes(semanas, ocupados_juan)
            self.aplicar_estilos_excel(ruta, df_c, df_j)

            if messagebox.askyesno("Listo", "El Excel se ha generado correctamente.\n¿Deseas abrirlo ahora?"):
                if sys.platform == "darwin":
                    subprocess.call(["open", ruta])
                elif sys.platform == "win32":
                    os.startfile(ruta)
                else:
                    subprocess.call(["xdg-open", ruta])

        except Exception as e:
            traceback.print_exc()
            messagebox.showerror("Error", f"No se pudo guardar: {e}")
        finally:
            self.btn.config(state="normal", text="Generar y Guardar Excel")
            self.root.config(cursor="")


if __name__ == "__main__":
    root = tk.Tk()
    AppTurnosNativa(root)
    root.mainloop()
