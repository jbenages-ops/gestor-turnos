import re
from collections import Counter
from datetime import datetime, timedelta

import pytest

import generador_turnos as gt

DIAS = gt.DIAS_SEMANA


@pytest.fixture
def app():
    return object.__new__(gt.AppTurnosNativa)


def _semanas(n, inicio=datetime(2026, 1, 5)):
    """n semanas consecutivas empezando en un lunes."""
    out = []
    curr = inicio
    for _ in range(n):
        out.append((curr, curr + timedelta(days=6)))
        curr += timedelta(days=7)
    return out


def _personas_de(celda):
    """Nombres entre paréntesis de una celda ('Tarea (Nombre)\n| ...')."""
    if celda == "-":
        return []
    return re.findall(r"\(([^)]+)\)", celda)


# ---------------------------------------------------------------------- #
#  _calcular_semanas                                                      #
# ---------------------------------------------------------------------- #

def test_calcular_semanas_empieza_en_lunes_y_cubre_rango(app):
    semanas = app._calcular_semanas(datetime(2026, 1, 7), datetime(2026, 2, 1))
    assert semanas[0][0] == datetime(2026, 1, 5)          # lunes previo al inicio
    assert all(f_ini.weekday() == 0 for f_ini, _ in semanas)
    assert all((f_fin - f_ini).days == 6 for f_ini, f_fin in semanas)
    assert semanas[-1][0] <= datetime(2026, 2, 1) <= semanas[-1][1]


def test_calcular_semanas_rango_de_una_semana(app):
    semanas = app._calcular_semanas(datetime(2026, 1, 5), datetime(2026, 1, 11))
    assert len(semanas) == 1


# ---------------------------------------------------------------------- #
#  generar_juan                                                           #
# ---------------------------------------------------------------------- #

def test_juan_respeta_disponibilidad(app):
    df, _ = app.generar_juan(_semanas(30))
    for _, row in df.iterrows():
        for dia in DIAS:
            for persona in _personas_de(row[dia]):
                assert dia in gt.DIAS_DISPONIBLES[persona], (
                    f"{persona} asignada el {dia}, pero no está disponible"
                )


def test_juan_asigna_las_tres_estancias_cada_semana(app):
    df, _ = app.generar_juan(_semanas(30))
    for _, row in df.iterrows():
        contenido = " ".join(row[d] for d in DIAS)
        for estancia in gt.ESTANCIAS_JUAN:
            assert estancia in contenido


def test_juan_usa_el_jueves(app):
    """Regresión: el jueves quedaba siempre vacío (columna muerta)."""
    df, _ = app.generar_juan(_semanas(30))
    assert (df["Jueves"] != "-").any()


def test_juan_estancias_rotan_entre_todo_el_staff(app):
    df, _ = app.generar_juan(_semanas(30))
    conteo = Counter()
    for _, row in df.iterrows():
        for dia in DIAS:
            conteo.update(_personas_de(row[dia]))
    assert set(conteo) == set(gt.STAFF_JUAN)


def test_juan_ocupados_coincide_con_la_tabla(app):
    df, ocupados = app.generar_juan(_semanas(10))
    assert len(ocupados) == len(df)
    for i, (_, row) in enumerate(df.iterrows()):
        for dia in DIAS:
            assert set(_personas_de(row[dia])) == set(ocupados[i][dia])


# ---------------------------------------------------------------------- #
#  generar_comunes                                                        #
# ---------------------------------------------------------------------- #

def test_comunes_cubre_los_dias_de_sus_bloques(app):
    df = app.generar_comunes(_semanas(9))
    dias_cubiertos = {d for bloque in gt.BLOQUES_COMUNES for d in (bloque[0], bloque[2])}
    for dia in DIAS:
        assert dia in df.columns
        if dia in dias_cubiertos:
            assert (df[dia] != "-").all()


def test_comunes_reparto_equilibrado(app):
    df = app.generar_comunes(_semanas(9))
    conteo = Counter()
    for _, row in df.iterrows():
        for dia in DIAS:
            conteo.update(_personas_de(row[dia]))
    assert len(set(conteo.values())) == 1, f"Reparto desigual: {dict(conteo)}"
    assert set(conteo) == set(gt.STAFF_COMUNES)


def test_comunes_evita_conflicto_con_zonas_de_juan(app):
    """Regresión: Lina acababa con tarea en ambas hojas el mismo día."""
    semanas = _semanas(60)
    df_j, ocupados = app.generar_juan(semanas)
    df_c = app.generar_comunes(semanas, ocupados)
    compartidos = set(gt.STAFF_JUAN) & set(gt.STAFF_COMUNES)
    assert compartidos, "el test presupone al menos una persona en ambos equipos"
    for i in range(len(semanas)):
        for dia in DIAS:
            en_juan = set(_personas_de(df_j.iloc[i][dia]))
            en_comunes = set(_personas_de(df_c.iloc[i][dia]))
            solapados = en_juan & en_comunes & compartidos
            assert not solapados, (
                f"Semana {i}, {dia}: {solapados} con tarea en ambas hojas"
            )


def test_comunes_sin_ocupados_funciona(app):
    df = app.generar_comunes(_semanas(3))
    assert len(df) == 3


# ---------------------------------------------------------------------- #
#  aplicar_estilos_excel                                                  #
# ---------------------------------------------------------------------- #

def test_excel_se_genera_con_dos_hojas(app, tmp_path):
    import openpyxl

    semanas = _semanas(4)
    df_j, ocupados = app.generar_juan(semanas)
    df_c = app.generar_comunes(semanas, ocupados)

    ruta = tmp_path / "turnos.xlsx"
    app.aplicar_estilos_excel(str(ruta), df_c, df_j)

    wb = openpyxl.load_workbook(ruta)
    assert wb.sheetnames == ["Zonas Comunes", "Zonas de Juan"]
    for ws in wb:
        assert ws.freeze_panes == "B2"
        assert ws.max_row == len(semanas) + 1
